#!/usr/bin/env python3
"""State persistence for 3-statements-ultra (v1.1).

Migrates from `_State` sheet inside xlsx to external state.json sidecar.

Sidecar path: alongside xlsx, named `<xlsx_stem>.state.json`.
  /path/to/ymm_3stmt.xlsx       -> /path/to/ymm_3stmt.state.json

Backwards-compatible migration:
  - load_state() first checks for state.json sidecar
  - If missing, falls back to parsing legacy _State sheet from xlsx
  - First migration auto-writes the sidecar (no manual conversion needed)
  - _State sheet is NOT auto-deleted (manual cleanup at user discretion)

JSON_FIELDS (RAW_MAP, ASM_MAP, etc.) are stored as native JSON objects/arrays
in the sidecar — no more `json.loads(state["RAW_MAP"])` boilerplate downstream.

Performance: state.json read ~50ms; legacy openpyxl _State read 1-3s on large
xlsx. Hooks that fire on every bash command (H2/H3) benefit ~20-50× speedup.
"""
from __future__ import annotations
import json
import os
import sys
from pathlib import Path
from typing import Optional, Union

# NOTE: do NOT re-wrap sys.stderr at import time — callers (hooks, scripts) already
# wrap their own; re-wrapping breaks the original wrapper and causes "lost sys.stderr".


SCHEMA_VERSION = 1

# Fields that were JSON-encoded strings in the legacy _State sheet but should
# be native types in state.json.
JSON_FIELDS = frozenset([
    "RAW_MAP", "ASM_MAP",
    "BS_COL_MAP", "CF_COL_MAP", "REV_BUILD_COL_MAP",
    "KEY_CELLS_IS", "KEY_CELLS_BS", "KEY_CELLS_CF",
    "HIST_YEARS", "FCST_YEARS",
    "DATA_SOURCES",
    "REV_BUILD_MAP", "REV_BUILD_SEGMENTS",
])


def state_json_path(xlsx_path: Union[str, Path]) -> Path:
    """Return the state.json sidecar path for a given xlsx."""
    p = Path(xlsx_path)
    return p.parent / (p.stem + ".state.json")


def _parse_legacy_state_sheet(xlsx_path: Union[str, Path]) -> dict:
    """Parse legacy `_State` sheet (KEY: VALUE per cell A) into a dict.
    Values for JSON_FIELDS are parsed from string to native types.
    File handle is always closed via try/finally (Codex v1.1 INFO fix)."""
    try:
        import openpyxl
    except ImportError:
        return {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=True)
    except Exception:
        return {}
    try:
        if "_State" not in wb.sheetnames:
            return {}
        state = {}
        for row in wb["_State"].iter_rows(values_only=True):
            if row and row[0] and isinstance(row[0], str) and ": " in row[0]:
                k, v = row[0].split(": ", 1)
                k = k.strip()
                v = v.strip()
                if k in JSON_FIELDS:
                    try:
                        v = json.loads(v)
                    except (json.JSONDecodeError, ValueError):
                        pass  # keep raw string if parse fails — caller can detect
                state[k] = v
        return state
    finally:
        try:
            wb.close()
        except Exception:
            pass


def load_state(xlsx_path: Union[str, Path]) -> dict:
    """Load state for an xlsx.

    1. If state.json sidecar exists AND xlsx hasn't been modified since the
       sidecar was written, read the sidecar (fast path).
    2. Else fall back to parsing legacy _State sheet from xlsx — this also
       covers the case where xlsx was edited after the sidecar (stale sidecar).
       The fresh parse auto-rewrites the sidecar.
    3. If neither has data, return empty dict.

    Always returns a dict (never None)."""
    xlsx_p = Path(xlsx_path)
    sidecar = state_json_path(xlsx_path)
    # Fast path: sidecar exists AND is newer than xlsx
    if sidecar.exists() and xlsx_p.exists():
        try:
            xlsx_mtime = xlsx_p.stat().st_mtime
            sidecar_mtime = sidecar.stat().st_mtime
            if sidecar_mtime >= xlsx_mtime:
                data = json.loads(sidecar.read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    return data
        except (json.JSONDecodeError, OSError, ValueError):
            sys.stderr.write(
                f"[state_io] WARNING: {sidecar.name} corrupt; falling back to _State sheet\n"
            )
    # Slow path: no sidecar, corrupt, or xlsx newer than sidecar — re-parse from _State sheet
    state = _parse_legacy_state_sheet(xlsx_path)
    if state:
        try:
            save_state(xlsx_path, state)
        except Exception as e:
            sys.stderr.write(f"[state_io] WARNING: sidecar save failed: {e}\n")
        return state
    # No _State sheet either — try the sidecar even if older (better than nothing)
    if sidecar.exists():
        try:
            data = json.loads(sidecar.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
        except (json.JSONDecodeError, OSError, ValueError):
            pass
    return {}


def save_state(xlsx_path: Union[str, Path], state: dict) -> None:
    """Write state dict to state.json sidecar atomically (tmp + rename)."""
    sidecar = state_json_path(xlsx_path)
    state = dict(state)  # don't mutate caller's dict
    state.setdefault("schema_version", SCHEMA_VERSION)
    sidecar.parent.mkdir(parents=True, exist_ok=True)
    tmp = sidecar.with_suffix(".json.tmp")
    payload = json.dumps(state, ensure_ascii=False, indent=2, sort_keys=False)
    tmp.write_text(payload, encoding="utf-8")
    os.replace(tmp, sidecar)


def write_key(xlsx_path: Union[str, Path], key: str, value) -> None:
    """Atomic update of a single key. Use this from gate scripts to write
    GATE_<X>_PASSED markers etc."""
    state = load_state(xlsx_path)
    state[key] = value
    save_state(xlsx_path, state)


def normalize_revenue_build(value) -> str:
    """Coerce REVENUE_BUILD state value to canonical "TRUE"/"FALSE"/"INVALID"/"UNSET".

    Accepts:
      - None / missing key → "UNSET"
      - bool True → "TRUE"
      - bool False → "FALSE"
      - str (case-insensitive) in {"TRUE", "FALSE"} → uppercased
      - anything else (str "YES"/"1"/"0"/int/float/list/...) → "INVALID"

    Callers must handle each case explicitly. Do NOT use truthiness — boolean True
    and string "TRUE" must both produce "TRUE"; string "YES" must NOT (it's a
    user error worth surfacing as INVALID rather than silently routing as FALSE
    or crashing on `.upper()`).
    """
    if value is None:
        return "UNSET"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, str):
        s = value.strip().upper()
        if s in ("TRUE", "FALSE"):
            return s
        return "INVALID"
    return "INVALID"


def has_legacy_state_sheet(xlsx_path: Union[str, Path]) -> bool:
    """True if the xlsx still contains a `_State` sheet (post-migration cleanup hint)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=True)
        present = "_State" in wb.sheetnames
        wb.close()
        return present
    except Exception:
        return False


def summary(xlsx_path: Union[str, Path]) -> str:
    """Human-readable one-line summary of current state."""
    state = load_state(xlsx_path)
    sidecar = state_json_path(xlsx_path)
    if not state:
        return f"[state_io] no state for {Path(xlsx_path).name}"
    has_legacy = has_legacy_state_sheet(xlsx_path)
    return (f"[state_io] {sidecar.name}: {len(state)} keys; "
            f"legacy _State sheet still present: {has_legacy}")


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="state.json sidecar tool")
    ap.add_argument("xlsx")
    ap.add_argument("--summary", action="store_true")
    ap.add_argument("--get", help="print value of key")
    ap.add_argument("--set", nargs=2, metavar=("KEY", "VALUE"),
        help="set KEY to VALUE (writes sidecar)")
    args = ap.parse_args()
    if args.summary:
        print(summary(args.xlsx))
    elif args.get:
        st = load_state(args.xlsx)
        print(st.get(args.get, ""))
    elif args.set:
        write_key(args.xlsx, args.set[0], args.set[1])
        print(f"set {args.set[0]} = {args.set[1]}")
    else:
        print(json.dumps(load_state(args.xlsx), ensure_ascii=False, indent=2))
