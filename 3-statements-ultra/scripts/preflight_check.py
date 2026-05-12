#!/usr/bin/env python3
"""
preflight_check.py — SESSION A end-of-phase gate

Runs after Raw_Info + Assumptions are built. Verifies the model is set up
correctly before SESSION B can start writing IS.

Severity: BLOCKER (exit 2) | WARNING (exit 1) | PASS (exit 0)

PF checks:
  PF-1  UNIT key in _State                       BLOCKER  (silent 100x error)
  PF-2  IS_GRANULARITY key in _State             BLOCKER  (COL_MAP structure)
  PF-3  DATA_SOURCES key with at least 1 source  BLOCKER  (provenance)
  PF-4  Raw_Info non-empty cell ratio >= 95%     BLOCKER  (downstream NaN)
  PF-5  NLM_0B_DONE flag set in _State           WARNING  (mgmt commentary)
  PF-6  YTD conversion done (Quarterly + Sina)   WARNING  (Q2=H1-Q1)
  PF-7  RAW_MAP / ASM_MAP spot-check passes      BLOCKER  (row stale)
  PF-8  NCI ratio set if hist NCI != 0           BLOCKER  (forecast NCI=0 bug)

Usage:
  python preflight_check.py --xlsx <path> [--json _preflight_report.json]
"""
from __future__ import annotations
import sys
import io
import json
import argparse
import re
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl not installed. Run: pip install openpyxl\n")
    sys.exit(127)

sys.path.insert(0, str(Path(__file__).resolve().parent))
import state_io  # v1.1


@dataclass
class Finding:
    pf: str
    severity: str
    msg: str = ""
    fix_hint: str = ""

    def __str__(self) -> str:
        return f"[{self.severity:7s}] {self.pf}: {self.msg}"


VALID_UNITS_RE = re.compile(
    r'(m|mn|百万|千万|千|万|百萬|千萬)\s*(RMB|CNY|USD|HKD|EUR|JPY|元)',
    re.IGNORECASE,
)
VALID_GRANULARITIES = {"Quarterly", "Semi-annual", "Annual"}


def read_state(xlsx_path) -> dict:
    """v1.1: load state.json sidecar (auto-migrates from _State sheet)."""
    return state_io.load_state(xlsx_path)


def pf1_unit(state) -> Finding:
    val = state.get("UNIT", "")
    if not val:
        return Finding("PF-1", "BLOCKER",
            "_State missing UNIT key",
            fix_hint="write_state_key(wb, 'UNIT', '百万元 RMB')")
    if not VALID_UNITS_RE.search(val):
        return Finding("PF-1", "BLOCKER",
            f"UNIT='{val}' not a recognized format",
            fix_hint="Use '百万元 RMB' / 'm USD' / '千万元 CNY' etc.")
    return Finding("PF-1", "PASS", f"UNIT='{val}'")


def pf2_granularity(state) -> Finding:
    val = state.get("IS_GRANULARITY", "")
    if not val:
        return Finding("PF-2", "BLOCKER",
            "_State missing IS_GRANULARITY key",
            fix_hint="write_state_key(wb, 'IS_GRANULARITY', 'Quarterly')")
    if val not in VALID_GRANULARITIES:
        return Finding("PF-2", "BLOCKER",
            f"IS_GRANULARITY='{val}' not in {VALID_GRANULARITIES}")
    return Finding("PF-2", "PASS", f"IS_GRANULARITY='{val}'")


def pf3_data_sources(state) -> Finding:
    val = state.get("DATA_SOURCES", "")
    if not val:
        return Finding("PF-3", "BLOCKER",
            "_State missing DATA_SOURCES key",
            fix_hint="DATA_SOURCES: [NLM, EXCEL, WEB] or similar")
    return Finding("PF-3", "PASS", f"DATA_SOURCES={val}")


def pf4_raw_info_nonempty(wb) -> Finding:
    if "Raw_Info" not in wb.sheetnames:
        return Finding("PF-4", "BLOCKER", "Raw_Info sheet does not exist")
    ws = wb["Raw_Info"]
    if ws.max_row < 5:
        return Finding("PF-4", "BLOCKER",
            f"Raw_Info only has {ws.max_row} rows — likely empty")
    label_rows = 0
    populated = 0
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 200),
                             min_col=1, max_col=min(ws.max_column, 20)):
        label = row[0].value
        if label and isinstance(label, str) and label.strip():
            label_rows += 1
            if any(c.value is not None for c in row[1:]):
                populated += 1
    if label_rows == 0:
        return Finding("PF-4", "BLOCKER", "Raw_Info has no label rows in col A")
    ratio = populated / label_rows
    if ratio < 0.95:
        return Finding("PF-4", "BLOCKER",
            f"Raw_Info populated ratio {ratio:.0%} ({populated}/{label_rows}) below 95%",
            fix_hint="Fill in missing rows or remove unused labels")
    return Finding("PF-4", "PASS", f"Raw_Info populated {ratio:.0%} ({populated}/{label_rows})")


def pf5_nlm_0b(state) -> Finding:
    val = (state.get("NLM_0B_DONE", "") or "").upper()
    if val in {"TRUE", "YES", "1"}:
        return Finding("PF-5", "PASS", "NLM_0B_DONE=TRUE")
    return Finding("PF-5", "WARNING",
        "NLM_0B_DONE not set — mgmt commentary / operational KPI not confirmed extracted",
        fix_hint="Ask NLM the 0B operational questions, then write_state_key(wb, 'NLM_0B_DONE', 'TRUE')")


def pf6_ytd_conversion(state, wb) -> Finding:
    granularity = state.get("IS_GRANULARITY", "")
    sources = state.get("DATA_SOURCES", "")
    if granularity != "Quarterly":
        return Finding("PF-6", "PASS", "non-Quarterly granularity, YTD check skipped")
    # v1.1: DATA_SOURCES may be list (state.json) or string (legacy); normalize for substring check
    sources_text = (" ".join(sources) if isinstance(sources, list) else str(sources)).lower()
    raw_data_form = str(state.get("RAW_DATA_FORM", "")).lower()
    if "sina" not in sources_text and "ytd" not in raw_data_form:
        return Finding("PF-6", "PASS", "no YTD source flagged")
    flag = (state.get("YTD_CONVERTED", "") or "").upper()
    if flag in {"TRUE", "YES", "1"}:
        return Finding("PF-6", "PASS", "YTD_CONVERTED=TRUE")
    return Finding("PF-6", "WARNING",
        "YTD source detected but YTD_CONVERTED not set — verify Q2=H1-Q1, Q3=Q3_YTD-H1",
        fix_hint="After conversion: write_state_key(wb, 'YTD_CONVERTED', 'TRUE')")


def pf7_map_spotcheck(state, wb) -> Finding:
    errors = []
    raw_map = state.get("RAW_MAP")
    asm_map = state.get("ASM_MAP")
    if raw_map is None:
        return Finding("PF-7", "BLOCKER", "state.json missing RAW_MAP")
    if asm_map is None:
        return Finding("PF-7", "BLOCKER", "state.json missing ASM_MAP")
    # v1.1: native types from sidecar; legacy callers may still pass strings
    if isinstance(raw_map, str):
        try:
            raw_map = json.loads(raw_map)
        except (json.JSONDecodeError, ValueError) as e:
            errors.append(f"RAW_MAP parse failed: {e}")
            raw_map = None
    if isinstance(asm_map, str):
        try:
            asm_map = json.loads(asm_map)
        except (json.JSONDecodeError, ValueError) as e:
            errors.append(f"ASM_MAP parse failed: {e}")
            asm_map = None
    if raw_map is not None and not isinstance(raw_map, dict):
        errors.append(f"RAW_MAP must be a JSON object/dict (got {type(raw_map).__name__})")
    elif raw_map is not None and "Raw_Info" in wb.sheetnames:
        ws = wb["Raw_Info"]
        for label, row in list(raw_map.items())[:3]:
            try:
                actual = ws.cell(int(row), 1).value
            except (TypeError, ValueError):
                errors.append(f"RAW_MAP['{label}'] row value {row!r} is not an integer")
                continue
            if actual != label:
                errors.append(f"RAW_MAP['{label}']={row} but Raw_Info!A{row}={actual!r}")
    if asm_map is not None and not isinstance(asm_map, dict):
        errors.append(f"ASM_MAP must be a JSON object/dict (got {type(asm_map).__name__})")
    elif asm_map is not None and "Assumptions" in wb.sheetnames:
        ws = wb["Assumptions"]
        for label, row in list(asm_map.items())[:3]:
            try:
                actual = ws.cell(int(row), 1).value
            except (TypeError, ValueError):
                errors.append(f"ASM_MAP['{label}'] row value {row!r} is not an integer")
                continue
            if actual != label:
                errors.append(f"ASM_MAP['{label}']={row} but Assumptions!A{row}={actual!r}")
    if errors:
        return Finding("PF-7", "BLOCKER",
            "Map spot-check failed: " + "; ".join(errors),
            fix_hint="Rebuild RAW_MAP/ASM_MAP from current Raw_Info/Assumptions labels")
    return Finding("PF-7", "PASS", "RAW_MAP + ASM_MAP spot-checks passed")


def pf8_nci_ratio(state, wb) -> Finding:
    """v1.1: RAW_MAP / ASM_MAP are native dicts from state.json (or legacy strings)."""
    raw_map = state.get("RAW_MAP")
    asm_map = state.get("ASM_MAP")
    if not raw_map or not asm_map:
        return Finding("PF-8", "PASS", "skipped — RAW_MAP/ASM_MAP not ready")
    # Accept either native dict (v1.1 state.json) or JSON string (legacy)
    if isinstance(raw_map, str):
        try:
            raw_map = json.loads(raw_map)
        except (json.JSONDecodeError, ValueError):
            return Finding("PF-8", "PASS", "skipped — RAW_MAP parse failed")
    if isinstance(asm_map, str):
        try:
            asm_map = json.loads(asm_map)
        except (json.JSONDecodeError, ValueError):
            return Finding("PF-8", "PASS", "skipped — ASM_MAP parse failed")
    if not isinstance(raw_map, dict) or not isinstance(asm_map, dict):
        return Finding("PF-8", "BLOCKER",
            f"RAW_MAP/ASM_MAP must be JSON objects/dicts "
            f"(got {type(raw_map).__name__}/{type(asm_map).__name__})",
            fix_hint="Rebuild RAW_MAP and ASM_MAP as JSON objects")
    nci_keys = [k for k in raw_map if "NCI" in k.upper() or "minority" in k.lower()]
    if not nci_keys:
        return Finding("PF-8", "PASS", "no NCI in Raw_Info — skipped")
    if "Raw_Info" not in wb.sheetnames:
        return Finding("PF-8", "PASS", "Raw_Info not found")
    ws = wb["Raw_Info"]
    has_hist_nci = False
    for label in nci_keys:
        try:
            row_num = int(raw_map[label])
        except (TypeError, ValueError):
            return Finding("PF-8", "BLOCKER",
                f"RAW_MAP['{label}'] row value {raw_map[label]!r} is not an integer",
                fix_hint="RAW_MAP values must be integer row numbers")
        if row_num < 1:
            return Finding("PF-8", "BLOCKER",
                f"RAW_MAP['{label}'] row value {row_num} is < 1 (openpyxl rejects)",
                fix_hint="RAW_MAP row numbers must be ≥ 1")
        try:
            for col in range(2, min(ws.max_column or 1, 15)):
                v = ws.cell(row_num, col).value
                if isinstance(v, (int, float)) and abs(v) > 0.01:
                    has_hist_nci = True
                    break
        except (ValueError, IndexError) as e:
            return Finding("PF-8", "BLOCKER",
                f"openpyxl rejected cell access at row {row_num} for '{label}': {e}",
                fix_hint="Verify RAW_MAP row numbers are valid Raw_Info row indices")
        if has_hist_nci:
            break
    if not has_hist_nci:
        return Finding("PF-8", "PASS", "historical NCI = 0, ratio not needed")
    asm_nci_keys = [k for k in asm_map if "NCI" in k.upper()]
    if not asm_nci_keys:
        return Finding("PF-8", "BLOCKER",
            "Historical NCI ≠ 0 but Assumptions has no 'NCI Ratio' row",
            fix_hint="Add Assumptions row 'NCI Ratio' before SESSION B")
    return Finding("PF-8", "PASS", f"NCI rows in Assumptions: {asm_nci_keys}")


def main():
    ap = argparse.ArgumentParser(description="SESSION A preflight gate")
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--json", default=None)
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.stderr.write(f"file not found: {xlsx_path}\n")
        sys.exit(127)

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    state = read_state(xlsx_path)
    if not state:
        sys.stderr.write(
            "[preflight] state.json sidecar missing and no legacy _State sheet — "
            "SESSION A not started\n"
        )
        sys.exit(2)

    findings = [
        pf1_unit(state),
        pf2_granularity(state),
        pf3_data_sources(state),
        pf4_raw_info_nonempty(wb),
        pf5_nlm_0b(state),
        pf6_ytd_conversion(state, wb),
        pf7_map_spotcheck(state, wb),
        pf8_nci_ratio(state, wb),
    ]

    summary = {"BLOCKER": 0, "WARNING": 0, "PASS": 0}
    for f in findings:
        summary[f.severity] = summary.get(f.severity, 0) + 1
    exit_code = 2 if summary["BLOCKER"] else (1 if summary["WARNING"] else 0)

    report = {
        "skill": "3-statements-ultra",
        "gate": "preflight",
        "xlsx": str(xlsx_path.resolve()),
        "ts": datetime.now().isoformat(timespec="seconds"),
        "summary": summary,
        "exit_code": exit_code,
        "findings": [asdict(f) for f in findings],
    }
    if args.json:
        Path(args.json).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        sys.stderr.write(f"[preflight] report: {args.json}\n")

    print(f"=== Preflight (SESSION A gate) ===")
    print(f"xlsx : {xlsx_path.name}")
    print(f"summary : BLOCKER={summary['BLOCKER']}  WARNING={summary['WARNING']}  PASS={summary['PASS']}")
    print()
    for f in findings:
        print(f"  {f}")
        if f.severity != "PASS" and f.fix_hint:
            print(f"      fix: {f.fix_hint}")
    print()
    print(f"exit code: {exit_code}")
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
