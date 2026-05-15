#!/usr/bin/env python3
"""
qc_suite.py — 3-Statements model QC gate

Severity: BLOCKER (exit 2) | WARNING (exit 1) | PASS (exit 0)

Modes:
  --smoke   Run only standalone format QCs (no _State sheet required).
            Works on any xlsx, e.g. completed reference models.
  --full    Run all 19 QCs (requires _State + RAW_MAP + ASM_MAP + KEY_CELLS_*).
            Used at end of SESSION E by per_session_gate.py.

Usage:
  python qc_suite.py --xlsx <path> --smoke
  python qc_suite.py --xlsx <path> --full --tabs IS,BS,CF --json _qc_report.json
"""
from __future__ import annotations
import sys
import io
import json
import argparse
import re
from dataclasses import dataclass, asdict, field
from datetime import datetime
from pathlib import Path
from typing import Optional, Callable

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl not installed. Run: pip install openpyxl\n")
    sys.exit(127)

# v1.1: state moved out of xlsx into sidecar JSON
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    import state_io
except ImportError:
    state_io = None


@dataclass
class Finding:
    qc: str
    severity: str
    sheet: Optional[str] = None
    cell: Optional[str] = None
    msg: str = ""
    fix_hint: str = ""

    def __str__(self) -> str:
        loc = f"{self.sheet}!{self.cell}" if self.sheet and self.cell else (self.sheet or "")
        return f"[{self.severity:7s}] {self.qc} {loc}: {self.msg}"


FINANCIAL_TAB_PATTERNS = [
    r"^IS$", r"^BS$", r"^CF$", r"^CFS$", r"^Assumptions$", r"^Revenue_Build$",
    r"IS\s*&?\s*drivers?", r"Balance\s*Sheet", r"Income\s*Statement",
    r"Cash\s*Flow", r"Historical\s+(IS|BS|CFS|CF)", r"Revenue\s*Build",
]

UNIT_PATTERN = re.compile(
    r"(?:Unit|单位)\s*[:：]?\s*"
    r"(?:m|mn|百万|千万|千|万|百萬|千萬)"
    r".{0,30}?"
    r"(?:RMB|CNY|USD|HKD|EUR|JPY|US\$|HK\$|￥|\$|元)",
    re.IGNORECASE,
)

ACCOUNTING_FMT_PATTERNS = [
    re.compile(r'#,##0'),
    re.compile(r'\*\s*#,##0'),
]

ALLOWED_FONTS = {"Book Antiqua", "等线", "Arial", "微软雅黑", "宋体", "Calibri"}
PREFERRED_FONTS = {"Book Antiqua", "等线"}

INPUT_BLUE_HEX = (0x00, 0x70, 0xC0)


def is_financial_tab(name: str) -> bool:
    for pat in FINANCIAL_TAB_PATTERNS:
        if re.search(pat, name, re.IGNORECASE):
            return True
    return False


def hex_distance(rgb_str: str, target: tuple) -> int:
    if not rgb_str or not isinstance(rgb_str, str):
        return 999
    s = rgb_str.upper().lstrip("FF")
    if len(s) != 6:
        return 999
    try:
        r, g, b = int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)
        return abs(r - target[0]) + abs(g - target[1]) + abs(b - target[2])
    except ValueError:
        return 999


def safe_rgb(color) -> Optional[str]:
    if color is None:
        return None
    try:
        rgb = color.rgb
        if isinstance(rgb, str):
            return rgb
    except (AttributeError, TypeError):
        pass
    return None


def qc8_gridlines(wb, tabs, xlsx_path=None) -> list[Finding]:
    out = []
    for tab in tabs:
        ws = wb[tab]
        if ws.sheet_view.showGridLines:
            out.append(Finding("QC-8", "WARNING", sheet=tab,
                msg="Gridlines enabled — disable for clean print/review",
                fix_hint="ws.sheet_view.showGridLines = False"))
        else:
            out.append(Finding("QC-8", "PASS", sheet=tab, msg="Gridlines disabled"))
    return out


def qc14_a1_unit(wb, tabs, xlsx_path=None) -> list[Finding]:
    out = []
    for tab in tabs:
        ws = wb[tab]
        a1 = ws.cell(1, 1).value
        if a1 is None or not isinstance(a1, str):
            out.append(Finding("QC-14", "BLOCKER", sheet=tab, cell="A1",
                msg="A1 empty or non-string; financial tab requires unit declaration",
                fix_hint='ws["A1"] = "Unit: m RMB"'))
            continue
        if UNIT_PATTERN.search(a1):
            out.append(Finding("QC-14", "PASS", sheet=tab, cell="A1",
                msg=f"A1 OK: '{a1[:80]}'"))
        else:
            out.append(Finding("QC-14", "BLOCKER", sheet=tab, cell="A1",
                msg=f"A1 does not match unit pattern: '{a1[:80]}'",
                fix_hint='Expected format: "Unit: m RMB" / "Unit: m USD" / "单位: 百万 RMB"'))
    return out


def qc15_font(wb, tabs, xlsx_path=None) -> list[Finding]:
    out = []
    for tab in tabs:
        ws = wb[tab]
        total = 0
        ok = 0
        preferred = 0
        max_r = min(ws.max_row or 1, 100)
        max_c = min(ws.max_column or 1, 40)
        for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
            for cell in row:
                if cell.value is None:
                    continue
                total += 1
                fn = cell.font.name
                if fn in ALLOWED_FONTS:
                    ok += 1
                if fn in PREFERRED_FONTS:
                    preferred += 1
        if total == 0:
            continue
        ratio = ok / total
        pref_ratio = preferred / total
        if ratio >= 0.90:
            out.append(Finding("QC-15", "PASS", sheet=tab,
                msg=f"Font compliance {ratio:.0%} ({ok}/{total}); preferred {pref_ratio:.0%}"))
        else:
            out.append(Finding("QC-15", "WARNING", sheet=tab,
                msg=f"Font compliance {ratio:.0%} ({ok}/{total}) below 90%",
                fix_hint=f"Use Book Antiqua (preferred) or 等线 for Chinese fallback"))
    return out


def qc17_number_format(wb, tabs, xlsx_path=None) -> list[Finding]:
    out = []
    for tab in tabs:
        ws = wb[tab]
        considered = 0
        compliant = 0
        max_r = min(ws.max_row or 1, 100)
        max_c = min(ws.max_column or 1, 40)
        for row in ws.iter_rows(min_row=2, max_row=max_r, min_col=2, max_col=max_c):
            for cell in row:
                if cell.value is None:
                    continue
                v = cell.value
                is_numlike = (
                    isinstance(v, (int, float))
                    or (isinstance(v, str) and v.startswith("="))
                )
                if not is_numlike:
                    continue
                considered += 1
                fmt = cell.number_format or ""
                if any(p.search(fmt) for p in ACCOUNTING_FMT_PATTERNS):
                    compliant += 1
                elif "%" in fmt and fmt != "General":
                    compliant += 1
                elif fmt in ("0", "0.0", "0.00", "0.000"):
                    compliant += 1
        if considered == 0:
            continue
        ratio = compliant / considered
        if ratio >= 0.70:
            out.append(Finding("QC-17", "PASS", sheet=tab,
                msg=f"Number format compliance {ratio:.0%} ({compliant}/{considered})"))
        else:
            out.append(Finding("QC-17", "WARNING", sheet=tab,
                msg=f"Number format compliance {ratio:.0%} below 70% ({compliant}/{considered})",
                fix_hint='Accounting: "_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"??_ ;_ @_ "  | Pct: "0.0%"'))
    return out


def qc18_zoom(wb, tabs, xlsx_path=None) -> list[Finding]:
    out = []
    for tab in tabs:
        ws = wb[tab]
        z = ws.sheet_view.zoomScale
        if z is None:
            out.append(Finding("QC-18", "PASS", sheet=tab, msg="Zoom unset (Excel default 100%)"))
            continue
        if 60 <= z <= 90:
            out.append(Finding("QC-18", "PASS", sheet=tab, msg=f"Zoom={z}%"))
        else:
            out.append(Finding("QC-18", "WARNING", sheet=tab,
                msg=f"Zoom={z}% outside recommended [60, 90]",
                fix_hint="ws.sheet_view.zoomScale = 70"))
    return out


def qc19_freeze(wb, tabs, xlsx_path=None) -> list[Finding]:
    out = []
    for tab in tabs:
        ws = wb[tab]
        if ws.freeze_panes:
            out.append(Finding("QC-19", "PASS", sheet=tab,
                msg=f"Freeze panes at {ws.freeze_panes}"))
        else:
            out.append(Finding("QC-19", "WARNING", sheet=tab,
                msg="No frozen panes — header rows/cols will scroll out of view",
                fix_hint='ws.freeze_panes = "B3" or appropriate anchor'))
    return out


def qc2_hardcode_smoke(wb, tabs, xlsx_path=None) -> list[Finding]:
    """Smoke version of QC-2: scan all numeric cells in IS/BS/CF.
    Without _State, we can't tell hist vs forecast cols, so we report counts only.
    Full version (requires _State) is enforced separately."""
    out = []
    for tab in tabs:
        if not re.search(r"^(IS|BS|CF|CFS)$|IS\s*&", tab, re.IGNORECASE):
            continue
        ws = wb[tab]
        numeric_cells = 0
        formula_cells = 0
        max_r = min(ws.max_row or 1, 200)
        max_c = min(ws.max_column or 1, 60)
        for row in ws.iter_rows(min_row=2, max_row=max_r, min_col=2, max_col=max_c):
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells += 1
                elif isinstance(cell.value, (int, float)):
                    numeric_cells += 1
        if numeric_cells + formula_cells == 0:
            continue
        fr = formula_cells / (numeric_cells + formula_cells)
        if fr < 0.50:
            out.append(Finding("QC-2", "WARNING", sheet=tab,
                msg=f"Formula ratio {fr:.0%} suspiciously low ({formula_cells} formulas vs {numeric_cells} numeric) — possible Rule Zero violations",
                fix_hint="Open tab in Excel and inspect cells with raw numbers in forecast columns"))
        else:
            out.append(Finding("QC-2", "PASS", sheet=tab,
                msg=f"Formula ratio {fr:.0%} ({formula_cells} formulas, {numeric_cells} numeric)"))
    return out


# =========================================================================
# Full-mode helpers: load _State context + value lookup utilities.
# =========================================================================

def get_full_context(wb, xlsx_path=None) -> Optional[dict]:
    """Load state via state_io (sidecar JSON; auto-migrates from _State sheet).
    Returns None if required keys missing — caller should emit a 'skipped' PASS.

    v1.1: state.json is the canonical source; JSON_FIELDS are native types
    (no more json.loads downstream)."""
    if state_io is None or xlsx_path is None:
        return None
    try:
        state = state_io.load_state(xlsx_path)
    except Exception:
        return None
    if not state:
        return None
    needed = ["RAW_MAP", "ASM_MAP", "BS_COL_MAP", "KEY_CELLS_IS",
              "KEY_CELLS_BS", "KEY_CELLS_CF", "FCST_YEARS", "HIST_YEARS"]
    if not all(k in state for k in needed):
        return None
    raw_map = state.get("RAW_MAP")
    asm_map = state.get("ASM_MAP")
    if not isinstance(raw_map, dict) or not isinstance(asm_map, dict):
        return None
    col_map = state.get("BS_COL_MAP")
    if not isinstance(col_map, dict):
        return None
    ctx = {
        "state": state,
        "raw_map": raw_map,
        "asm_map": asm_map,
        "col_map": col_map,
        "key_is": state.get("KEY_CELLS_IS") if isinstance(state.get("KEY_CELLS_IS"), dict) else {},
        "key_bs": state.get("KEY_CELLS_BS") if isinstance(state.get("KEY_CELLS_BS"), dict) else {},
        "key_cf": state.get("KEY_CELLS_CF") if isinstance(state.get("KEY_CELLS_CF"), dict) else {},
        "fcst_years": state.get("FCST_YEARS") if isinstance(state.get("FCST_YEARS"), list) else [],
        "hist_years": state.get("HIST_YEARS") if isinstance(state.get("HIST_YEARS"), list) else [],
    }
    cf_col_map = state.get("CF_COL_MAP")
    if isinstance(cf_col_map, dict):
        ctx["cf_col_map"] = cf_col_map
    return ctx


def _col_idx(col_map, yr) -> Optional[int]:
    if yr not in col_map:
        return None
    try:
        return openpyxl.utils.column_index_from_string(col_map[yr])
    except (ValueError, KeyError):
        return None


def _rv(ws_raw, raw_map, col_map, label, yr):
    """Raw_Info value lookup. Returns 0 on any failure."""
    if label not in raw_map:
        return 0
    ci = _col_idx(col_map, yr)
    if ci is None:
        return 0
    try:
        v = ws_raw.cell(int(raw_map[label]), ci).value
        return v if isinstance(v, (int, float)) else 0
    except (ValueError, TypeError):
        return 0


def _av(ws_asm, asm_map, fcst_years, label, yr):
    """Assumptions value lookup. Returns 0 on any failure."""
    if label not in asm_map or yr not in fcst_years:
        return 0
    try:
        fcst_col_idx = fcst_years.index(yr) + 2
        v = ws_asm.cell(int(asm_map[label]), fcst_col_idx).value
        return v if isinstance(v, (int, float)) else 0
    except (ValueError, TypeError):
        return 0


# =========================================================================
# Full-mode QCs (ported from references/session-e.md L225-403, plus 4 new).
# All take signature (wb, tabs, xlsx_path) — xlsx_path optional, needed only
# for QCs that reopen workbook with data_only=True.
# Each gracefully skips if expected schema (asm_map labels, etc.) is missing.
# =========================================================================

def qc1_numeric_reconciliation(wb, tabs, xlsx_path=None) -> list[Finding]:
    """Read BS_CHECK_row + CF_CHECK_row computed values from data_only workbook.
    The model's spreadsheet has =Total_Assets - Total_LE and =Ending_Cash - BS_Cash
    formulas in these rows. We verify they evaluate to 0 across all periods.

    Codex review note: prior implementation (ported from session-e.md) was
    tautological — it solved Others as the balance plug then checked balance,
    which by construction returned 0 regardless of model defects. This version
    reads actual computed cell values instead."""
    ctx = get_full_context(wb, xlsx_path)
    if ctx is None:
        return [Finding("QC-1", "PASS", msg="skipped — _State context not ready")]
    key_bs, key_cf, col_map = ctx["key_bs"], ctx["key_cf"], ctx["col_map"]
    if "BS_CHECK_row" not in key_bs or "CF_CHECK_row" not in key_cf:
        return [Finding("QC-1", "PASS",
            msg="skipped — BS_CHECK_row/CF_CHECK_row not defined in KEY_CELLS_*")]
    if not xlsx_path or not Path(xlsx_path).exists():
        return [Finding("QC-1", "PASS",
            msg="skipped — xlsx_path required for data_only read",
            fix_hint="Caller should pass --xlsx pointing to readable file")]
    try:
        wb_v = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return [Finding("QC-1", "WARNING",
            msg=f"data_only reload failed: {type(e).__name__}: {e}",
            fix_hint="Open xlsx in Excel and save once to populate cached values")]
    if "BS" not in wb_v.sheetnames or "CF" not in wb_v.sheetnames:
        return [Finding("QC-1", "PASS", msg="skipped — BS/CF missing in data_only workbook")]
    ws_bs_v, ws_cf_v = wb_v["BS"], wb_v["CF"]
    out = []
    checked = 0
    for yr, col_letter in col_map.items():
        try:
            ci = openpyxl.utils.column_index_from_string(col_letter)
        except (ValueError, KeyError):
            continue
        try:
            bs_check = ws_bs_v.cell(int(key_bs["BS_CHECK_row"]), ci).value
            cf_check = ws_cf_v.cell(int(key_cf["CF_CHECK_row"]), ci).value
        except (ValueError, TypeError):
            continue
        # Require BOTH cells to have cached values; a half-computed pair is not a valid check
        if bs_check is None or cf_check is None:
            continue
        checked += 1
        if isinstance(bs_check, (int, float)) and abs(bs_check) > 0.01:
            out.append(Finding("QC-1", "BLOCKER", sheet="BS",
                cell=f"{col_letter}{key_bs['BS_CHECK_row']}",
                msg=f"BS_CHECK {yr} = {bs_check:.4f} (expected 0)",
                fix_hint="BS does not balance — inspect R3 Others plug, NCI roll, or Assumptions divergence"))
        if isinstance(cf_check, (int, float)) and abs(cf_check) > 0.01:
            out.append(Finding("QC-1", "BLOCKER", sheet="CF",
                cell=f"{col_letter}{key_cf['CF_CHECK_row']}",
                msg=f"CF_CHECK {yr} = {cf_check:.4f} (expected 0)",
                fix_hint="CF Ending Cash diverges from BS Cash — verify back-fill order"))
    if checked == 0:
        out.append(Finding("QC-1", "WARNING",
            msg="skipped — all BS_CHECK/CF_CHECK cells empty (xlsx not yet calculated by Excel?)",
            fix_hint="Open xlsx in Excel and save once to compute formula values"))
    elif not any(f.severity == "BLOCKER" for f in out):
        out.append(Finding("QC-1", "PASS",
            msg=f"BS_CHECK + CF_CHECK = 0 across {checked} periods (data_only read)"))
    return out


def qc2_hardcode_full(wb, tabs, xlsx_path=None) -> list[Finding]:
    """Forecast cells in IS/BS/CF must be Excel formula strings, not int/float."""
    ctx = get_full_context(wb, xlsx_path)
    if ctx is None:
        return [Finding("QC-2", "PASS", msg="skipped — _State context not ready")]
    key_is, key_bs, key_cf, col_map, fcst_years = (
        ctx["key_is"], ctx["key_bs"], ctx["key_cf"], ctx["col_map"], ctx["fcst_years"])
    fcst_cols = [_col_idx(col_map, yr) for yr in fcst_years if _col_idx(col_map, yr)]
    targets = []
    if "IS" in wb.sheetnames and "NCI_row" in key_is:
        targets.append(("IS", wb["IS"], range(2, int(key_is["NCI_row"]) + 1)))
    if "BS" in wb.sheetnames and "BS_CHECK_row" in key_bs:
        targets.append(("BS", wb["BS"], range(2, int(key_bs["BS_CHECK_row"]) + 1)))
    if "CF" in wb.sheetnames and "CF_CHECK_row" in key_cf:
        targets.append(("CF", wb["CF"], range(2, int(key_cf["CF_CHECK_row"]) + 1)))
    out = []
    hits = 0
    for tab_name, ws, row_range in targets:
        for col in fcst_cols:
            for row in row_range:
                v = ws.cell(row, col).value
                if isinstance(v, (int, float)):
                    out.append(Finding("QC-2", "BLOCKER", sheet=tab_name,
                        cell=f"{openpyxl.utils.get_column_letter(col)}{row}",
                        msg=f"hardcoded value {v} in forecast cell",
                        fix_hint="Replace with Excel formula string starting with '='"))
                    hits += 1
                    if hits >= 20:
                        out.append(Finding("QC-2", "BLOCKER",
                            msg=f"...truncated; {hits}+ total hardcodes in forecast cells"))
                        return out
    if not out:
        out.append(Finding("QC-2", "PASS",
            msg=f"forecast cells across IS/BS/CF clean ({len(fcst_cols)} cols × {len(targets)} tabs)"))
    return out


def qc3_others_cfo(wb, tabs, xlsx_path=None) -> list[Finding]:
    """|Others/CFO| <= 15% PASS; 15-30% WARNING; > 30% BLOCKER."""
    ctx = get_full_context(wb, xlsx_path)
    if ctx is None or "cf_col_map" not in ctx or "CF" not in wb.sheetnames:
        return [Finding("QC-3", "PASS", msg="skipped — CF_COL_MAP or CF sheet not ready")]
    key_cf = ctx["key_cf"]
    if "Others_row" not in key_cf or "CFO_row" not in key_cf:
        return [Finding("QC-3", "PASS", msg="skipped — Others_row/CFO_row not in KEY_CELLS_CF")]
    if not xlsx_path or not Path(xlsx_path).exists():
        return [Finding("QC-3", "PASS",
            msg="skipped — xlsx_path required for data_only read",
            fix_hint="Optional check; caller may pass --xlsx for completeness")]
    try:
        wb_v = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return [Finding("QC-3", "WARNING", msg=f"data_only reload failed: {e}")]
    if "CF" not in wb_v.sheetnames:
        return [Finding("QC-3", "PASS", msg="skipped — CF sheet not in data_only workbook")]
    ws_cf_v = wb_v["CF"]
    cf_col_map = ctx["cf_col_map"]
    qc3_cols = {k: v for k, v in cf_col_map.items()
                if not any(k.endswith(s) for s in ("Q1", "Q2", "Q3", "Q4", "H1", "H2"))}
    out = []
    for yr, col in qc3_cols.items():
        ci = _col_idx({yr: col}, yr)
        if ci is None:
            continue
        others = ws_cf_v.cell(int(key_cf["Others_row"]), ci).value or 0
        cfo = ws_cf_v.cell(int(key_cf["CFO_row"]), ci).value or 0
        if not isinstance(others, (int, float)) or not isinstance(cfo, (int, float)):
            continue
        if cfo == 0:
            continue
        ratio = abs(others / cfo)
        if ratio > 0.30:
            out.append(Finding("QC-3", "BLOCKER", sheet="CF",
                msg=f"|Others/CFO| {ratio:.1%} > 30% for {yr} — material CF line missing",
                fix_hint="Identify and explicitly model the missing line (ROU amort, JV income, DTA, etc.)"))
        elif ratio > 0.15:
            out.append(Finding("QC-3", "WARNING", sheet="CF",
                msg=f"|Others/CFO| {ratio:.1%} > 15% for {yr}",
                fix_hint="Investigate CF line completeness"))
    if not any(f.severity in ("BLOCKER", "WARNING") for f in out):
        out.append(Finding("QC-3", "PASS",
            msg=f"|Others/CFO| ≤ 15% across {len(qc3_cols)} annual periods"))
    return out


_CF_REF_RE = re.compile(r"^\s*=\s*'?CF'?\s*!", re.IGNORECASE)


def qc4_bs_cash_backfill(wb, tabs, xlsx_path=None) -> list[Finding]:
    """BS Cash cells must be formulas referencing the CF sheet (back-filled from CF Ending Cash).
    Accepts =CF!, ='CF'!, lowercase, with/without leading whitespace (Excel quirks)."""
    ctx = get_full_context(wb, xlsx_path)
    if ctx is None or "BS" not in wb.sheetnames:
        return [Finding("QC-4", "PASS", msg="skipped — _State or BS not ready")]
    key_bs, col_map = ctx["key_bs"], ctx["col_map"]
    if "Cash_row" not in key_bs:
        return [Finding("QC-4", "PASS", msg="skipped — Cash_row not in KEY_CELLS_BS")]
    ws_bs = wb["BS"]
    out = []
    for yr, col_letter in col_map.items():
        try:
            ci = openpyxl.utils.column_index_from_string(col_letter)
            val = ws_bs.cell(int(key_bs["Cash_row"]), ci).value
        except (ValueError, KeyError):
            continue
        is_cf_ref = isinstance(val, str) and bool(_CF_REF_RE.search(val))
        if not is_cf_ref:
            out.append(Finding("QC-4", "BLOCKER", sheet="BS",
                cell=f"{col_letter}{key_bs['Cash_row']}",
                msg=f"BS Cash {yr} not linked to CF (value: {val!r})",
                fix_hint=f"Set ws_bs[{col_letter}{key_bs['Cash_row']}].value = '=CF!<EndingCash>'"))
    if not out:
        out.append(Finding("QC-4", "PASS",
            msg=f"all {len(col_map)} BS Cash cells linked to CF Ending Cash"))
    return out


_IS_REF_RE = re.compile(r"(?<![A-Za-z0-9])'?IS'?\s*!", re.IGNORECASE)


def qc5_nci_continuity(wb, tabs, xlsx_path=None) -> list[Finding]:
    """BS NCI forecast cells must contain reference to IS (rolling from Attr_to_NCI).
    Accepts IS! or 'IS'! (Excel may quote sheet names)."""
    ctx = get_full_context(wb, xlsx_path)
    if ctx is None or "BS" not in wb.sheetnames:
        return [Finding("QC-5", "PASS", msg="skipped — _State or BS not ready")]
    key_bs, col_map, fcst_years = ctx["key_bs"], ctx["col_map"], ctx["fcst_years"]
    if "NCI_row" not in key_bs:
        return [Finding("QC-5", "PASS", msg="skipped — NCI_row not in KEY_CELLS_BS")]
    ws_bs = wb["BS"]
    out = []
    for yr in fcst_years:
        ci = _col_idx(col_map, yr)
        if ci is None:
            continue
        val = ws_bs.cell(int(key_bs["NCI_row"]), ci).value
        has_is_ref = isinstance(val, str) and bool(_IS_REF_RE.search(val))
        if not has_is_ref:
            out.append(Finding("QC-5", "BLOCKER", sheet="BS",
                cell=f"{col_map[yr]}{key_bs['NCI_row']}",
                msg=f"BS NCI {yr} not rolling from IS (value: {val!r})",
                fix_hint="Forecast NCI must reference IS Attr_to_NCI: =Prior_NCI + IS!<Attr_to_NCI>"))
    if not out:
        out.append(Finding("QC-5", "PASS",
            msg=f"NCI rolling formula confirmed in all {len(fcst_years)} forecast years"))
    return out


def qc6_formula_integrity(wb, tabs, xlsx_path=None) -> list[Finding]:
    """Random spot-check: IS hist rows -> =Raw_Info!; BS fcst rows -> any formula."""
    import random
    ctx = get_full_context(wb, xlsx_path)
    if ctx is None:
        return [Finding("QC-6", "PASS", msg="skipped — _State not ready")]
    key_is, key_bs = ctx["key_is"], ctx["key_bs"]
    out = []
    if "IS" in wb.sheetnames and "NCI_row" in key_is and "first_hist_col" in key_is:
        ws_is = wb["IS"]
        nrows = int(key_is["NCI_row"]) - 1
        if nrows >= 2:
            sample = random.sample(list(range(2, int(key_is["NCI_row"]) + 1)), min(5, nrows))
            for row in sample:
                val = ws_is.cell(row, int(key_is["first_hist_col"])).value
                if val is not None and not (isinstance(val, str) and val.startswith("=Raw_Info!")):
                    out.append(Finding("QC-6", "BLOCKER", sheet="IS",
                        cell=f"{openpyxl.utils.get_column_letter(int(key_is['first_hist_col']))}{row}",
                        msg=f"IS hist row {row} not linked to Raw_Info (value: {val!r})",
                        fix_hint="Historical cells must be formula =Raw_Info!<col><row>"))
    if "BS" in wb.sheetnames and "BS_CHECK_row" in key_bs and "first_fcst_col" in key_bs:
        ws_bs = wb["BS"]
        nrows = int(key_bs["BS_CHECK_row"]) - 2
        if nrows >= 2:
            sample = random.sample(list(range(2, int(key_bs["BS_CHECK_row"]))), min(5, nrows))
            for row in sample:
                val = ws_bs.cell(row, int(key_bs["first_fcst_col"])).value
                if val is not None and val != 0 and not (isinstance(val, str) and val.startswith("=")):
                    out.append(Finding("QC-6", "BLOCKER", sheet="BS",
                        cell=f"{openpyxl.utils.get_column_letter(int(key_bs['first_fcst_col']))}{row}",
                        msg=f"BS fcst row {row} not a formula (value: {val!r})",
                        fix_hint="Forecast cells must be Excel formula strings"))
    if not out:
        out.append(Finding("QC-6", "PASS", msg="formula integrity spot-check passed"))
    return out


def qc7_state_keys(wb, tabs, xlsx_path=None) -> list[Finding]:
    """state.json (or legacy _State sheet) must contain all 14 required keys."""
    required = ["FILE_HASH", "RAW_MAP", "ASM_MAP", "IS_GRANULARITY", "BS_COL_MAP",
                "CF_COL_MAP", "KEY_CELLS_IS", "KEY_CELLS_BS", "KEY_CELLS_CF",
                "BS_CASH_STATUS", "HIST_YEARS", "FCST_YEARS", "UNIT", "GAAP"]
    if state_io is None or xlsx_path is None:
        # Fallback: read _State sheet directly (legacy path)
        if "_State" not in wb.sheetnames:
            return [Finding("QC-7", "BLOCKER", msg="state.json + _State both missing")]
        state = {}
        for row in wb["_State"].iter_rows(values_only=True):
            if row and row[0] and isinstance(row[0], str) and ": " in row[0]:
                k, _ = row[0].split(": ", 1)
                state[k.strip()] = True
    else:
        state = state_io.load_state(xlsx_path)
        if not state:
            return [Finding("QC-7", "BLOCKER",
                msg=f"state.json sidecar missing and no legacy _State sheet for {Path(xlsx_path).name}")]
    missing = [k for k in required if k not in state]
    if missing:
        return [Finding("QC-7", "BLOCKER", sheet="state.json",
            msg=f"missing keys: {missing}",
            fix_hint="See references/gate-spec.md for the canonical 14-key list")]
    return [Finding("QC-7", "PASS", msg=f"all {len(required)} required state keys present")]


def qc9_summary_hardcode(wb, tabs, xlsx_path=None) -> list[Finding]:
    """Summary cells must be formulas linking to model — no hardcoded numbers."""
    if "Summary" not in wb.sheetnames:
        return [Finding("QC-9", "PASS", msg="skipped — Summary sheet not yet built")]
    ws = wb["Summary"]
    out = []
    hits = 0
    max_r = min(ws.max_row or 1, 100)
    max_c = min(ws.max_column or 1, 30)
    for row in ws.iter_rows(min_row=2, max_row=max_r, min_col=2, max_col=max_c):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                out.append(Finding("QC-9", "BLOCKER", sheet="Summary",
                    cell=cell.coordinate,
                    msg=f"hardcoded value {cell.value}",
                    fix_hint="Link to model cell via KEY_CELLS_* reference"))
                hits += 1
                if hits >= 10:
                    out.append(Finding("QC-9", "BLOCKER",
                        msg=f"...truncated; {hits}+ hardcodes in Summary"))
                    return out
    if not out:
        out.append(Finding("QC-9", "PASS", msg="Summary cells all formula-linked"))
    return out


def qc10_wc_days_sanity(wb, tabs, xlsx_path=None) -> list[Finding]:
    """Forecast AR/Inv/AP days vs 3yr historical avg — deviation >30% = WARNING."""
    ctx = get_full_context(wb, xlsx_path)
    if ctx is None:
        return [Finding("QC-10", "PASS", msg="skipped — _State not ready")]
    raw_map, asm_map, col_map = ctx["raw_map"], ctx["asm_map"], ctx["col_map"]
    fcst_years, hist_years = ctx["fcst_years"], ctx["hist_years"]
    ws_raw = wb["Raw_Info"] if "Raw_Info" in wb.sheetnames else None
    ws_asm = wb["Assumptions"] if "Assumptions" in wb.sheetnames else None
    if not ws_raw or not ws_asm:
        return [Finding("QC-10", "PASS", msg="skipped — Raw_Info/Assumptions missing")]

    out = []
    metrics = [
        ("AR Days", "BS: AR", "IS: Revenue"),
        ("Inv Days", "BS: Inventory", "IS: COGS"),
        ("AP Days", "BS: AP", "IS: COGS"),
    ]
    for asm_label, num_label, denom_label in metrics:
        if asm_label not in asm_map or num_label not in raw_map or denom_label not in raw_map:
            continue
        recent_hist = hist_years[-3:] if len(hist_years) >= 3 else hist_years
        hist_days = []
        for yr in recent_hist:
            num = _rv(ws_raw, raw_map, col_map, num_label, yr)
            denom = _rv(ws_raw, raw_map, col_map, denom_label, yr)
            if abs(denom) < 1:
                continue
            hist_days.append(abs(num) / abs(denom) * 365)
        if not hist_days:
            continue
        hist_avg = sum(hist_days) / len(hist_days)
        if hist_avg < 1:
            continue
        for yr in fcst_years:
            fcst_days = _av(ws_asm, asm_map, fcst_years, asm_label, yr)
            if fcst_days == 0:
                continue
            deviation = abs(fcst_days - hist_avg) / hist_avg
            if deviation > 0.30:
                out.append(Finding("QC-10", "WARNING", sheet="Assumptions",
                    msg=f"{asm_label} {yr}={fcst_days:.1f}d deviates {deviation:.0%} from 3yr hist avg {hist_avg:.1f}d",
                    fix_hint=f"Justify in notes or revise to within ±30% of {hist_avg:.1f}d"))
    if not out:
        out.append(Finding("QC-10", "PASS", msg="WC days within ±30% of 3yr historical avg"))
    return out


def qc11_segment_sum(wb, tabs, xlsx_path=None) -> list[Finding]:
    """Σ(Revenue segments) must equal Total Revenue across all periods."""
    ctx = get_full_context(wb, xlsx_path)
    if ctx is None:
        return [Finding("QC-11", "PASS", msg="skipped — _State not ready")]
    raw_map, col_map = ctx["raw_map"], ctx["col_map"]
    if "Raw_Info" not in wb.sheetnames:
        return [Finding("QC-11", "PASS", msg="skipped — Raw_Info missing")]
    ws_raw = wb["Raw_Info"]
    total_label = next((k for k in ["IS: Revenue", "IS: Total Revenue"] if k in raw_map), None)
    if not total_label:
        return [Finding("QC-11", "PASS", msg="skipped — no total revenue row")]
    segment_labels = [k for k in raw_map
                      if k.startswith("IS: Revenue ") and k != total_label
                      and "YoY" not in k and "%" not in k]
    if not segment_labels:
        return [Finding("QC-11", "PASS", msg="skipped — no segment rows (single-segment model)")]
    out = []
    all_years = ctx["hist_years"] + ctx["fcst_years"]
    for yr in all_years:
        seg_sum = sum(_rv(ws_raw, raw_map, col_map, lbl, yr) for lbl in segment_labels)
        total = _rv(ws_raw, raw_map, col_map, total_label, yr)
        if abs(total) < 1:
            continue
        rel_err = abs(seg_sum - total) / abs(total)
        if rel_err > 0.01:
            out.append(Finding("QC-11", "BLOCKER", sheet="Raw_Info",
                msg=f"Σ(segments)={seg_sum:.1f} ≠ Total {total:.1f} for {yr} ({len(segment_labels)} segs, err {rel_err:.1%})",
                fix_hint="Verify each segment uses independent source data; no double-counting"))
    if not out:
        out.append(Finding("QC-11", "PASS",
            msg=f"Σ(segments)=Total verified across all periods ({len(segment_labels)} segments)"))
    return out


def qc12_tax_rate(wb, tabs, xlsx_path=None) -> list[Finding]:
    """Forecast tax rate vs 3yr historical effective rate — deviation > 5pct = WARNING."""
    ctx = get_full_context(wb, xlsx_path)
    if ctx is None:
        return [Finding("QC-12", "PASS", msg="skipped — _State not ready")]
    raw_map, asm_map, col_map = ctx["raw_map"], ctx["asm_map"], ctx["col_map"]
    fcst_years, hist_years = ctx["fcst_years"], ctx["hist_years"]
    tax_label = next((k for k in ["IS: Tax", "IS: Income Tax", "IS: Tax Expense"] if k in raw_map), None)
    pbt_label = next((k for k in ["IS: PBT", "IS: Pre-Tax Income", "IS: Profit before Tax"] if k in raw_map), None)
    if not tax_label or not pbt_label or "Tax Rate %" not in asm_map:
        return [Finding("QC-12", "PASS", msg="skipped — tax/PBT/Tax Rate labels not in expected format")]
    if "Raw_Info" not in wb.sheetnames or "Assumptions" not in wb.sheetnames:
        return [Finding("QC-12", "PASS", msg="skipped — sheets missing")]
    ws_raw, ws_asm = wb["Raw_Info"], wb["Assumptions"]
    # Sign-agnostic: take |tax|/|pbt|. Works for both conventions:
    #   US/CN GAAP (tax negative when PBT positive) and IFRS (tax expense positive).
    # Filter out implausible rates (> 100% effective = data issue or non-recurring credit).
    hist_rates = []
    skipped_implausible = 0
    for yr in hist_years[-3:]:
        tax = _rv(ws_raw, raw_map, col_map, tax_label, yr)
        pbt = _rv(ws_raw, raw_map, col_map, pbt_label, yr)
        if abs(pbt) < 1:
            continue
        eff = abs(tax) / abs(pbt)
        if eff > 1.0:
            skipped_implausible += 1
            continue
        hist_rates.append(eff)
    if not hist_rates:
        msg = "skipped — no usable hist tax data"
        if skipped_implausible:
            msg += f" ({skipped_implausible} years with effective rate > 100% skipped)"
        return [Finding("QC-12", "PASS", msg=msg)]
    hist_avg = sum(hist_rates) / len(hist_rates)
    out = []
    for yr in fcst_years:
        fcst_rate = _av(ws_asm, asm_map, fcst_years, "Tax Rate %", yr)
        if fcst_rate == 0:
            continue
        if abs(fcst_rate - hist_avg) > 0.05:
            out.append(Finding("QC-12", "WARNING", sheet="Assumptions",
                msg=f"Tax Rate {yr}={fcst_rate:.1%} deviates >5pct from {len(hist_rates)}yr hist avg {hist_avg:.1%}",
                fix_hint="Verify HNTE/incentive expiry or revise forecast tax rate"))
    if not out:
        msg = f"forecast tax rates within ±5pct of hist avg {hist_avg:.1%}"
        if skipped_implausible:
            msg += f" ({skipped_implausible} years excluded as implausible)"
        out.append(Finding("QC-12", "PASS", msg=msg))
    return out


def qc13_others_historical_source(wb, tabs, xlsx_path=None) -> list[Finding]:
    """R3 Others in historical years should have Raw_Info source — not silently 0 as plug."""
    ctx = get_full_context(wb, xlsx_path)
    if ctx is None or "CF" not in wb.sheetnames:
        return [Finding("QC-13", "PASS", msg="skipped — _State or CF missing")]
    key_cf, col_map, hist_years, raw_map = (
        ctx["key_cf"], ctx["col_map"], ctx["hist_years"], ctx["raw_map"])
    if "Others_row" not in key_cf:
        return [Finding("QC-13", "PASS", msg="skipped — Others_row not in KEY_CELLS_CF")]
    ws_cf = wb["CF"]
    raw_others_exists = any(("Others" in k or "Other" in k) and k.startswith("CF:")
                             for k in raw_map)
    out = []
    for yr in hist_years:
        ci = _col_idx(col_map, yr)
        if ci is None:
            continue
        val = ws_cf.cell(int(key_cf["Others_row"]), ci).value
        is_formula = isinstance(val, str) and val.startswith("=")
        is_zero = val is None or val == 0
        if not is_formula and is_zero and not raw_others_exists:
            out.append(Finding("QC-13", "WARNING", sheet="CF",
                cell=f"{col_map[yr]}{key_cf['Others_row']}",
                msg=f"CF Others {yr}=0 with no Raw_Info source registered (silent plug)",
                fix_hint="Populate from disclosed Raw_Info value, or document why 0 is intended"))
    if not out:
        out.append(Finding("QC-13", "PASS", msg="CF Others historical source check passed"))
    return out


_REVENUE_BUILD_REF_RE = re.compile(r"^\s*=\s*'?Revenue_Build'?\s*!", re.IGNORECASE)

_QUARTER_RE = re.compile(r"^[1-4]Q", re.IGNORECASE)
_SUMMARY_RE = re.compile(r"^(FY|H[12]\b)", re.IGNORECASE)

def _is_summary_period(period_name: str) -> bool:
    """True if the period is a summary column (FY or H1/H2 semi-annual).
    Quarter columns (1Q/2Q/3Q/4Q) return False.
    Unknown patterns default to False (conservative: treated as Q-like → R12 strict)."""
    if not period_name:
        return False
    return bool(_SUMMARY_RE.match(period_name.strip()))

# Codex round-2 B3 refinement: aggregate-key filter must be precise. Using prefix-match
# on "total" was too broad — a legitimate segment named "Total Vehicles" (e.g., NEV vs ICE
# total) would be falsely excluded as aggregate. Match only well-defined aggregate labels:
#   - "Grand Total" / "Total Revenue" exact
#   - any label starting with "Reconciliation"
#   - any label ending with "Delta" / "Δ" / "Delta %"  (reconciliation deltas)
#   - "Check ..." / "Sum ..." prefixes
# Crucially: bare "Auto Total" / "Handset Total" / "Total Vehicles" segment keys DO NOT match.
_AGGREGATE_KEY_RE = re.compile(
    r"^(?:"
    r"grand[\s_]*total\b.*"       # Grand Total / Grand Total Revenue / Grand Total Auto / etc.
    r"|total[\s_]*revenue\b.*"    # Total Revenue (sum row)
    r"|reconciliation\b.*"        # Reconciliation Delta % / Reconciliation Check / etc.
    r"|.*(?:delta|δ)\s*%?"        # X Delta / X Δ / X Delta % (reconciliation deltas)
    r"|check\b.*"                 # Check rows
    r"|sum\b.*"                   # Sum rows
    r")$",
    re.IGNORECASE,
)


def qc20_revenue_build_integrity(wb, tabs, xlsx_path=None) -> list[Finding]:
    """R12 — Revenue_Build forecast cells must be formulas, never hardcoded.

    Scans Revenue_Build tab if REVENUE_BUILD=TRUE in state. Each forecast-column cell
    in the data range (excluding labels/headers/notes) must be a string starting with '='.
    Historical-column cells are allowed to be hardcoded (data source rows).

    Codex B1 fix: REVENUE_BUILD value normalized via state_io.normalize_revenue_build
    (handles bool/str/None uniformly; INVALID values surface as a BLOCKER, not silent skip).

    BLOCKER if any forecast cell holds an int/float instead of a formula string.
    """
    if state_io is None or xlsx_path is None:
        return [Finding("QC-20", "PASS", msg="skipped — state_io / xlsx_path required")]
    try:
        state = state_io.load_state(xlsx_path)
    except Exception:
        return [Finding("QC-20", "PASS", msg="skipped — state.json not loadable")]
    if not state:
        return [Finding("QC-20", "PASS", msg="skipped — state empty")]
    rb_norm = state_io.normalize_revenue_build(state.get("REVENUE_BUILD"))
    if rb_norm == "INVALID":
        return [Finding("QC-20", "BLOCKER",
            msg=f"REVENUE_BUILD={state.get('REVENUE_BUILD')!r} is invalid (expected TRUE/FALSE)",
            fix_hint="Re-run SESSION A Phase 2.5 to set REVENUE_BUILD to TRUE or FALSE")]
    if rb_norm != "TRUE":
        return [Finding("QC-20", "PASS",
            msg=f"skipped — REVENUE_BUILD={rb_norm}; Revenue_Build not in use")]
    if "Revenue_Build" not in wb.sheetnames:
        return [Finding("QC-20", "BLOCKER", sheet="Revenue_Build",
            msg="REVENUE_BUILD=TRUE in state but Revenue_Build tab missing",
            fix_hint="Build the Revenue_Build tab per references/session-a2.md")]
    ws = wb["Revenue_Build"]
    # Revenue_Build may have different column positions than IS (e.g., a retrofitted
    # model: FY2026E = col T in Revenue_Build but col X in IS). Use REV_BUILD_COL_MAP if
    # available; fall back to BS_COL_MAP for models where columns are aligned.
    rb_col_map = state.get("REV_BUILD_COL_MAP")
    if not isinstance(rb_col_map, dict) or not rb_col_map:
        rb_col_map = state.get("BS_COL_MAP") if isinstance(state.get("BS_COL_MAP"), dict) else {}
    fcst_years = state.get("FCST_YEARS") if isinstance(state.get("FCST_YEARS"), list) else []
    fcst_cols = [_col_idx(rb_col_map, yr) for yr in fcst_years if _col_idx(rb_col_map, yr)]
    if not fcst_cols:
        return [Finding("QC-20", "PASS",
            msg="skipped — no forecast columns resolved from REV_BUILD_COL_MAP/BS_COL_MAP + FCST_YEARS")]
    # Build reverse period map once (hoisted from row loop per Codex W1).
    _rb_reverse = {}
    for period, col_l in rb_col_map.items():
        ci = _col_idx(rb_col_map, period)
        if ci is not None:
            _rb_reverse[ci] = period
    q_fcst_cols = [c for c in fcst_cols if not _is_summary_period(_rb_reverse.get(c, ""))]
    # Scan all data rows in forecast columns.
    out = []
    hits = 0
    max_r = min(ws.max_row or 1, 200)
    for r in range(4, max_r + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        label_str = str(label).strip()
        if not label_str or label_str.startswith("•") or label_str.startswith("—"):
            continue
        is_section_hdr = (label_str.isupper() and len(label_str) > 5 and
                          not any(ws.cell(r, c).value is not None for c in fcst_cols))
        if is_section_hdr:
            continue
        if not any(ws.cell(r, c).value is not None for c in fcst_cols):
            continue
        # Quarterly: check which column types have data in this row
        summary_fcst_cols = [c for c in fcst_cols if c not in q_fcst_cols]
        has_any_q_data = any(ws.cell(r, c).value is not None for c in q_fcst_cols)
        has_any_summary_data = any(ws.cell(r, c).value is not None for c in summary_fcst_cols)
        for col in fcst_cols:
            v = ws.cell(r, col).value
            col_letter = openpyxl.utils.get_column_letter(col)
            period_name = _rb_reverse.get(col, "")
            is_summary = _is_summary_period(period_name)
            if v is None:
                if is_summary and has_any_summary_data:
                    out.append(Finding("QC-20", "BLOCKER", sheet="Revenue_Build",
                        cell=f"{col_letter}{r}",
                        msg=f"blank summary forecast cell in Revenue_Build '{label_str[:30]}' (must be formula)",
                        fix_hint="Write a formula: =SUM(Q1:Q4) or =Assumptions!<col><row>"))
                    hits += 1
                elif not is_summary and has_any_q_data:
                    out.append(Finding("QC-20", "BLOCKER", sheet="Revenue_Build",
                        cell=f"{col_letter}{r}",
                        msg=f"blank Q forecast cell in Revenue_Build '{label_str[:30]}' (must be formula)",
                        fix_hint="Write a formula: =prior_Q*(1+Assumptions!YoY) or =Vol*ASP"))
                    hits += 1
                # else: structural blank (Q-only row blank FY, or summary-only row blank Q) → structural blank → OK
            elif isinstance(v, (int, float)):
                out.append(Finding("QC-20", "BLOCKER", sheet="Revenue_Build",
                    cell=f"{col_letter}{r}",
                    msg=f"hardcoded value {v} in Revenue_Build forecast cell (R12 violation)",
                    fix_hint="Replace with formula string: =prior*(1+Assumptions!YoY) or =Vol*ASP/scale"))
                hits += 1
            elif isinstance(v, str) and not v.strip().startswith("="):
                out.append(Finding("QC-20", "BLOCKER", sheet="Revenue_Build",
                    cell=f"{col_letter}{r}",
                    msg=f"text constant '{v[:40]}' in forecast cell (R12: must be formula starting with '=')",
                    fix_hint="Replace with formula string starting with '='"))
                hits += 1
            # else: v is a formula string starting with "=" → OK
            if hits >= 20:
                out.append(Finding("QC-20", "BLOCKER",
                    msg=f"...truncated; {hits}+ non-formula cells in Revenue_Build forecast"))
                return out
    # Also verify REV_BUILD_MAP is populated
    rbm = state.get("REV_BUILD_MAP")
    if not rbm or not isinstance(rbm, dict) or len(rbm) == 0:
        out.append(Finding("QC-20", "BLOCKER",
            msg="REV_BUILD_MAP missing/empty in state.json",
            fix_hint="SESSION A2 Step 11 must write {segment_name: total_row} mapping"))
    if not out:
        out.append(Finding("QC-20", "PASS",
            msg=f"Revenue_Build forecast cells clean ({len(fcst_cols)} cols scanned, "
                f"REV_BUILD_MAP has {len(rbm) if isinstance(rbm, dict) else 0} entries)"))
    return out


def qc21_is_revenue_build_link(wb, tabs, xlsx_path=None) -> list[Finding]:
    """R12 — When REVENUE_BUILD=TRUE, IS forecast Revenue-by-segment cells must reference
    Revenue_Build, not Assumptions YoY directly.

    Codex B3 fix: segment-to-row mapping is read from explicit
    `KEY_CELLS_IS["Revenue_segment_rows"]` (set by SESSION B when IS is built).
    No substring/token heuristic — eliminates false positives on header rows containing
    segment-name fragments, and language-agnostic (works with 汽车 IS label vs "Auto"
    REV_BUILD_MAP key as long as SESSION B writes the explicit mapping).

    Skip / block matrix:
      REVENUE_BUILD unset/false   → PASS skip
      REVENUE_BUILD invalid       → BLOCKER (surface bad state)
      REVENUE_BUILD=TRUE, REV_BUILD_MAP empty → BLOCKER
      REVENUE_BUILD=TRUE, IS sheet absent → PASS skip (IS not built yet)
      REVENUE_BUILD=TRUE, IS exists, Revenue_segment_rows mapping absent → BLOCKER
      REVENUE_BUILD=TRUE, mapping present, any segment missing → BLOCKER
      Each mapped row × forecast col cell → must be string starting with =Revenue_Build!
    """
    if state_io is None or xlsx_path is None:
        return [Finding("QC-21", "PASS", msg="skipped — state_io / xlsx_path required")]
    try:
        state = state_io.load_state(xlsx_path)
    except Exception:
        return [Finding("QC-21", "PASS", msg="skipped — state.json not loadable")]
    if not state:
        return [Finding("QC-21", "PASS", msg="skipped — state empty")]
    rb_norm = state_io.normalize_revenue_build(state.get("REVENUE_BUILD"))
    if rb_norm == "INVALID":
        return [Finding("QC-21", "BLOCKER",
            msg=f"REVENUE_BUILD={state.get('REVENUE_BUILD')!r} is invalid (expected TRUE/FALSE)",
            fix_hint="Re-run SESSION A Phase 2.5 to set REVENUE_BUILD to TRUE or FALSE")]
    if rb_norm != "TRUE":
        return [Finding("QC-21", "PASS",
            msg=f"skipped — REVENUE_BUILD={rb_norm}; IS may use Assumptions directly")]
    if "IS" not in wb.sheetnames:
        return [Finding("QC-21", "PASS", msg="skipped — IS sheet not built yet")]
    rbm = state.get("REV_BUILD_MAP")
    if not isinstance(rbm, dict) or not rbm:
        return [Finding("QC-21", "BLOCKER",
            msg="REVENUE_BUILD=TRUE but REV_BUILD_MAP empty",
            fix_hint="Run SESSION A2 to populate REV_BUILD_MAP")]
    key_is = state.get("KEY_CELLS_IS") if isinstance(state.get("KEY_CELLS_IS"), dict) else {}
    seg_rows = key_is.get("Revenue_segment_rows")
    if not isinstance(seg_rows, dict) or not seg_rows:
        return [Finding("QC-21", "BLOCKER",
            msg="KEY_CELLS_IS['Revenue_segment_rows'] missing — SESSION B must register "
                "{segment_name: is_row_number} when building IS Revenue section",
            fix_hint="In SESSION B after writing IS Revenue rows, set "
                "key_cells_is['Revenue_segment_rows'] = {seg: row} and write KEY_CELLS_IS to state")]
    # Required segments derived from REV_BUILD_MAP, excluding aggregate/reconciliation keys.
    # Codex round-2 B3 fix: use anchored regex (not prefix-match) so legitimate segments
    # like "Total Vehicles" are NOT excluded just because they start with "total".
    required_segs = [k for k in rbm.keys() if not _AGGREGATE_KEY_RE.match(k.strip())]
    # Also strip trailing " Total" suffix for matching (REV_BUILD_MAP keys are "Auto Total"
    # whereas Revenue_segment_rows keys typically read "Auto").
    def _strip_total(s):
        return re.sub(r"\s+total$", "", s, flags=re.IGNORECASE).strip()
    missing = []
    resolved = {}  # {seg_label_in_rbm: is_row}
    for seg in required_segs:
        candidates = [seg, _strip_total(seg)]
        found_row = None
        for c in candidates:
            if c in seg_rows:
                found_row = seg_rows[c]
                break
        if found_row is None:
            missing.append(seg)
        else:
            try:
                resolved[seg] = int(found_row)
            except (ValueError, TypeError):
                missing.append(f"{seg} (row value not int: {found_row!r})")
    if missing:
        return [Finding("QC-21", "BLOCKER", sheet="IS",
            msg=f"Revenue_segment_rows missing entries for: {missing}",
            fix_hint="SESSION B must register every REV_BUILD_MAP non-aggregate segment "
                "in KEY_CELLS_IS['Revenue_segment_rows']")]
    col_map = state.get("BS_COL_MAP") if isinstance(state.get("BS_COL_MAP"), dict) else {}
    fcst_years = state.get("FCST_YEARS") if isinstance(state.get("FCST_YEARS"), list) else []
    fcst_cols = [_col_idx(col_map, yr) for yr in fcst_years if _col_idx(col_map, yr)]
    if not fcst_cols:
        return [Finding("QC-21", "PASS",
            msg="skipped — no forecast columns resolved")]
    # Build reverse map: col_idx → period_name, to detect Q vs summary columns.
    # Q cells must =Revenue_Build!* (R12 strict).
    # Summary cells (FY / H1 / H2) can be =SUM(Q1:Q4) — any formula accepted.
    _reverse_col_map = {}
    for period, col_l in col_map.items():
        ci = _col_idx(col_map, period)
        if ci is not None:
            _reverse_col_map[ci] = period
    ws_is = wb["IS"]
    out = []
    hits = 0
    checked = 0
    q_checked = 0
    fy_checked = 0
    for seg, is_row in resolved.items():
        for col in fcst_cols:
            cell = ws_is.cell(is_row, col)
            v = cell.value
            checked += 1
            col_letter = openpyxl.utils.get_column_letter(col)
            period_name = _reverse_col_map.get(col, "")
            is_summary = _is_summary_period(period_name)
            if is_summary:
                fy_checked += 1
            else:
                q_checked += 1
            # All forecast cells: blank and hardcode always BLOCKER
            if v is None:
                out.append(Finding("QC-21", "BLOCKER", sheet="IS",
                    cell=f"{col_letter}{is_row}",
                    msg=f"IS forecast Revenue ({seg}) is blank — must be formula (R12)",
                    fix_hint=f"Write =Revenue_Build!<col><seg_total_row> (Q col) or =SUM(Q1:Q4) (FY col)"))
                hits += 1
            elif isinstance(v, (int, float)):
                out.append(Finding("QC-21", "BLOCKER", sheet="IS",
                    cell=f"{col_letter}{is_row}",
                    msg=f"IS forecast Revenue ({seg}) is hardcoded value {v} — R12 violation",
                    fix_hint=f"Replace with =Revenue_Build!{col_letter}<seg_total_row>"))
                hits += 1
            elif isinstance(v, str):
                if is_summary:
                    # Summary columns (FY/H1/H2): accept ANY formula (typically =SUM(Q1:Q4)).
                    # The Q cells they sum are individually R12-checked.
                    if not v.strip().startswith("="):
                        out.append(Finding("QC-21", "BLOCKER", sheet="IS",
                            cell=f"{col_letter}{is_row}",
                            msg=f"IS forecast Revenue ({seg}) summary cell = '{v[:40]}' is text constant (must be formula)",
                            fix_hint="Summary cell should be =SUM(Q1:Q4) or =Revenue_Build!<col><row>"))
                        hits += 1
                else:
                    # Q columns: must reference Revenue_Build (strict R12)
                    if not _REVENUE_BUILD_REF_RE.match(v):
                        out.append(Finding("QC-21", "BLOCKER", sheet="IS",
                            cell=f"{col_letter}{is_row}",
                            msg=f"IS forecast Revenue ({seg}) Q cell = '{v[:60]}' does not reference Revenue_Build (R12)",
                            fix_hint=f"When REVENUE_BUILD=TRUE, Q forecast cell must start with '=Revenue_Build!'"))
                        hits += 1
            if hits >= 20:
                out.append(Finding("QC-21", "BLOCKER",
                    msg=f"...truncated; {hits}+ IS Revenue cells bypass Revenue_Build"))
                return out
    if not out:
        out.append(Finding("QC-21", "PASS",
            msg=f"IS forecast Revenue rows correctly reference Revenue_Build "
                f"({len(resolved)} segs × {len(fcst_cols)} fcst cols = {checked} cells: "
                f"{q_checked} Q-level R12-strict + {fy_checked} FY-level formula-any)"))
    return out


_INLINE_LABEL_RE = re.compile(r"^\s*(↳|->)\s+", re.UNICODE)


def qc22_inline_hist_completeness(wb, tabs, xlsx_path=None) -> list[Finding]:
    """Historical inline ↳ cells must be filled wherever a prior period exists.

    BLOCKER if a ↳ row has a blank historical cell but at least one earlier column
    in the SAME ↳ row already has a value (proving a prior exists and the ratio is
    computable). First-period blanks (no earlier data in the row at all) are OK.

    Scans Revenue_Build and IS tabs (wherever ↳ rows appear).
    """
    if state_io is None or xlsx_path is None:
        return [Finding("QC-22", "PASS", msg="skipped — state_io / xlsx_path required")]
    try:
        state = state_io.load_state(xlsx_path)
    except Exception:
        return [Finding("QC-22", "PASS", msg="skipped — state not loadable")]
    if not state:
        return [Finding("QC-22", "PASS", msg="skipped — state empty")]
    col_map = state.get("BS_COL_MAP") if isinstance(state.get("BS_COL_MAP"), dict) else {}
    hist_years = state.get("HIST_YEARS") if isinstance(state.get("HIST_YEARS"), list) else []
    if not hist_years or not col_map:
        return [Finding("QC-22", "PASS", msg="skipped — no HIST_YEARS or COL_MAP")]
    hist_cols = [_col_idx(col_map, yr) for yr in hist_years if _col_idx(col_map, yr)]
    hist_cols_sorted = sorted(hist_cols)
    if not hist_cols_sorted:
        return [Finding("QC-22", "PASS", msg="skipped — no historical columns resolved")]
    # Also check Revenue_Build with its own COL_MAP if present
    rb_col_map = state.get("REV_BUILD_COL_MAP")
    if not isinstance(rb_col_map, dict):
        rb_col_map = col_map
    rb_hist_cols = sorted([_col_idx(rb_col_map, yr) for yr in hist_years
                           if _col_idx(rb_col_map, yr)])
    scan_tabs = []
    if "Revenue_Build" in wb.sheetnames:
        scan_tabs.append(("Revenue_Build", wb["Revenue_Build"], rb_hist_cols))
    if "IS" in wb.sheetnames:
        scan_tabs.append(("IS", wb["IS"], hist_cols_sorted))
    if not scan_tabs:
        return [Finding("QC-22", "PASS", msg="skipped — no Revenue_Build or IS tab")]
    out = []
    hits = 0
    total_checked = 0
    for tab_name, ws, h_cols in scan_tabs:
        max_r = min(ws.max_row or 1, 200)
        for r in range(1, max_r + 1):
            label = ws.cell(r, 1).value
            if not isinstance(label, str):
                continue
            if not _INLINE_LABEL_RE.match(label):
                continue
            # This is a ↳ row. Determine if Q-active or FY/summary-only.
            # Build reverse map for period classification
            _rev = {}
            for period, col_l in (rb_col_map if tab_name == "Revenue_Build" else col_map).items():
                _ci = _col_idx(rb_col_map if tab_name == "Revenue_Build" else col_map, period)
                if _ci is not None:
                    _rev[_ci] = period
            q_hist = [c for c in h_cols if not _is_summary_period(_rev.get(c, ""))]
            has_any_hist_q = any(ws.cell(r, c).value is not None for c in q_hist)
            # Scan same-type columns only: if row has Q data → check Q+FY; if FY-only → check FY only
            row_vals = [(c, ws.cell(r, c).value) for c in h_cols]
            seen_any_value = False
            for c_idx, v in row_vals:
                is_summary_col = _is_summary_period(_rev.get(c_idx, ""))
                if v is not None:
                    seen_any_value = True
                elif seen_any_value:
                    # Blank after prior data exists. But skip structural blanks:
                    # - Q blank in FY-only row → structural → skip
                    if not is_summary_col and not has_any_hist_q:
                        total_checked += 1
                        continue
                    col_letter = openpyxl.utils.get_column_letter(c_idx)
                    out.append(Finding("QC-22", "BLOCKER", sheet=tab_name,
                        cell=f"{col_letter}{r}",
                        msg=f"historical inline row '{label.strip()[:35]}' has blank cell "
                            f"but prior period data exists (fill implied ratio)",
                        fix_hint="Compute implied ratio: =curr/prior-1 (YoY%) or =data/total (%)"))
                    hits += 1
                    if hits >= 15:
                        out.append(Finding("QC-22", "BLOCKER",
                            msg=f"...truncated; {hits}+ blank historical inline cells"))
                        return out
                total_checked += 1
    if not out:
        out.append(Finding("QC-22", "PASS",
            msg=f"historical inline rows complete ({total_checked} cells across {len(scan_tabs)} tabs)"))
    return out


SMOKE_QCS: list[Callable] = [
    qc8_gridlines,
    qc14_a1_unit,
    qc15_font,
    qc17_number_format,
    qc18_zoom,
    qc19_freeze,
    qc2_hardcode_smoke,
]

FULL_QCS: list[Callable] = [
    # Format checks (shared with smoke)
    qc8_gridlines, qc14_a1_unit, qc15_font, qc17_number_format, qc18_zoom, qc19_freeze,
    # Numeric / structural checks (full only)
    qc1_numeric_reconciliation,
    qc2_hardcode_full,
    qc3_others_cfo,
    qc4_bs_cash_backfill,
    qc5_nci_continuity,
    qc6_formula_integrity,
    qc7_state_keys,
    qc9_summary_hardcode,
    qc10_wc_days_sanity,
    qc11_segment_sum,
    qc12_tax_rate,
    qc13_others_historical_source,
    # R12 — Revenue_Build (conditional, auto-skips if REVENUE_BUILD!=TRUE)
    qc20_revenue_build_integrity,
    qc21_is_revenue_build_link,
    qc22_inline_hist_completeness,
]


def detect_tabs(wb, override: str) -> list[str]:
    if override == "auto":
        return [s for s in wb.sheetnames if is_financial_tab(s)]
    return [t.strip() for t in override.split(",") if t.strip() in wb.sheetnames]


def main():
    ap = argparse.ArgumentParser(description="3-Statements QC suite gate")
    ap.add_argument("--xlsx", required=True, help="path to model .xlsx")
    ap.add_argument("--tabs", default="auto",
        help="comma-separated tab names; 'auto' picks financial tabs by name")
    ap.add_argument("--smoke", action="store_true",
        help="run only standalone format QCs (no _State required)")
    ap.add_argument("--full", action="store_true",
        help="run all QCs (requires _State sheet)")
    ap.add_argument("--json", default=None, help="write JSON report to this path")
    args = ap.parse_args()

    if not args.smoke and not args.full:
        args.smoke = True

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.stderr.write(f"file not found: {xlsx_path}\n")
        sys.exit(127)

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    tabs = detect_tabs(wb, args.tabs)

    if not tabs:
        sys.stderr.write(f"[qc_suite] no financial tabs detected; sheets: {wb.sheetnames}\n")
        sys.exit(1)

    findings: list[Finding] = []
    qcs = SMOKE_QCS if args.smoke else FULL_QCS
    for fn in qcs:
        try:
            findings.extend(fn(wb, tabs, xlsx_path))
        except Exception as e:
            findings.append(Finding(
                qc=f"QC-{fn.__name__}",
                severity="WARNING",
                msg=f"check raised exception: {type(e).__name__}: {e}",
            ))

    summary = {"BLOCKER": 0, "WARNING": 0, "PASS": 0}
    for f in findings:
        summary[f.severity] = summary.get(f.severity, 0) + 1

    if summary["BLOCKER"] > 0:
        exit_code = 2
    elif summary["WARNING"] > 0:
        exit_code = 1
    else:
        exit_code = 0

    report = {
        "skill": "3-statements-ultra",
        "xlsx": str(xlsx_path.resolve()),
        "ts": datetime.now().isoformat(timespec="seconds"),
        "tabs_checked": tabs,
        "mode": "smoke" if args.smoke else "full",
        "summary": summary,
        "exit_code": exit_code,
        "findings": [asdict(f) for f in findings],
    }

    if args.json:
        json_path = Path(args.json)
        json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        sys.stderr.write(f"[qc_suite] report written: {json_path}\n")

    print(f"=== 3-Statements QC Suite ({'smoke' if args.smoke else 'full'}) ===")
    print(f"xlsx : {xlsx_path.name}")
    print(f"tabs : {', '.join(tabs)}")
    print(f"summary : BLOCKER={summary['BLOCKER']}  WARNING={summary['WARNING']}  PASS={summary['PASS']}")
    print()
    shown_blocker = [f for f in findings if f.severity == "BLOCKER"]
    shown_warning = [f for f in findings if f.severity == "WARNING"]
    if shown_blocker:
        print("--- BLOCKER ---")
        for f in shown_blocker:
            print(f"  {f}")
            if f.fix_hint:
                print(f"      fix: {f.fix_hint}")
        print()
    if shown_warning:
        print("--- WARNING ---")
        for f in shown_warning:
            print(f"  {f}")
        print()
    print(f"exit code: {exit_code}")
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
