#!/usr/bin/env python3
"""
PreToolUse Bash hook — 3-Statements _State Consistency Guard

Catches the most common silent failure after context compaction: continuing
to write to a 3-statements model whose _State sheet has stale RAW_MAP or
ASM_MAP (row numbers shifted vs Raw_Info / Assumptions labels). This results
in the next session writing to the wrong rows — silent garbage.

Fires only when bash command:
  (a) opens an .xlsx that matches 3-statements model naming, AND
  (b) command body references RAW_MAP / ASM_MAP / KEY_CELLS_* (i.e. about to use map)

Exit codes:
  0  no obvious inconsistency, or hook out of scope
  2  BLOCKER — RAW_MAP / ASM_MAP spot-check failed (file or _State is stale)
"""
import sys
import io
import json
import re
from pathlib import Path

try:
    sys.stdin  = io.TextIOWrapper(sys.stdin.buffer,  encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
except Exception:
    pass


def read_hook_input():
    try:
        return json.load(sys.stdin)
    except Exception:
        return {}


XLSX_PATH_PATTERN = re.compile(r'["\']([^"\']{0,400}\.xlsx)["\']')
THREE_STMT_HINT = re.compile(
    r'3statement|3-statement|three[_-]stmt'
    r'|RAW_MAP|ASM_MAP|KEY_CELLS_(IS|BS|CF)|BS_COL_MAP|CF_COL_MAP'
    r'|_State|Raw_Info|Assumptions',
    re.IGNORECASE,
)
MAP_USE_PATTERN = re.compile(
    r'RAW_MAP|ASM_MAP|raw_map\s*\[|asm_map\s*\[|json\.loads\s*\(\s*state\s*\[',
    re.IGNORECASE,
)


def extract_xlsx_path(cmd: str) -> str | None:
    for m in XLSX_PATH_PATTERN.finditer(cmd):
        p = m.group(1)
        if THREE_STMT_HINT.search(p) or any(
            t in cmd[:m.start()] + cmd[m.end():m.end()+200].lower()
            for t in ("3statement", "three_stmt")
        ):
            return p
        if Path(p).exists():
            return p
    return None


def spot_check_map(ws_target, map_dict: dict, label_col: int = 1, n: int = 3) -> tuple[bool, str]:
    """Return (ok, msg). Picks first `n` entries from map_dict, verifies
    ws_target.cell(row, label_col).value == label."""
    if not map_dict:
        return True, "empty map (skipped)"
    items = list(map_dict.items())[:n]
    for label, row in items:
        try:
            actual = ws_target.cell(int(row), label_col).value
        except Exception as e:
            return False, f"failed to read cell row={row} col={label_col}: {e}"
        if actual != label:
            return False, f"expected '{label}' at row {row}, got {actual!r}"
    return True, f"verified {len(items)} entries"


def main():
    hook = read_hook_input()
    cmd = (hook.get("tool_input", {}) or {}).get("command", "") or ""

    if not THREE_STMT_HINT.search(cmd):
        sys.exit(0)
    if not MAP_USE_PATTERN.search(cmd):
        sys.exit(0)

    xlsx = extract_xlsx_path(cmd)
    if not xlsx:
        sys.exit(0)

    p = Path(xlsx)
    if not p.exists() or not p.is_file():
        sys.exit(0)

    # v1.1: state via state_io sidecar (auto-migrates from _State sheet, much faster).
    state_io_dir = Path.home() / ".claude" / "skills" / "3-statements-ultra" / "scripts"
    if state_io_dir.exists():
        sys.path.insert(0, str(state_io_dir))
    try:
        import state_io  # type: ignore
        state = state_io.load_state(p)
    except Exception:
        sys.exit(0)
    if not state:
        sys.exit(0)

    raw_map = state.get("RAW_MAP")
    asm_map = state.get("ASM_MAP")
    if raw_map is None and asm_map is None:
        sys.exit(0)

    # v1.1: maps may be native dicts (state.json) or JSON strings (legacy fallback)
    if isinstance(raw_map, str):
        try:
            raw_map = json.loads(raw_map)
        except json.JSONDecodeError:
            raw_map = None
    if isinstance(asm_map, str):
        try:
            asm_map = json.loads(asm_map)
        except json.JSONDecodeError:
            asm_map = None

    # Only open xlsx if we have maps to spot-check
    if not (isinstance(raw_map, dict) or isinstance(asm_map, dict)):
        sys.exit(0)

    try:
        import openpyxl
    except ImportError:
        sys.exit(0)
    try:
        wb = openpyxl.load_workbook(p, data_only=False, read_only=True)
    except Exception:
        sys.exit(0)

    errors = []
    if isinstance(raw_map, dict) and "Raw_Info" in wb.sheetnames:
        ok, msg = spot_check_map(wb["Raw_Info"], raw_map)
        if not ok:
            errors.append(f"RAW_MAP spot-check: {msg}")
    if isinstance(asm_map, dict) and "Assumptions" in wb.sheetnames:
        ok, msg = spot_check_map(wb["Assumptions"], asm_map)
        if not ok:
            errors.append(f"ASM_MAP spot-check: {msg}")

    if errors:
        sys.stderr.write(
            "[3stmt-state-guard] BLOCKER: _State inconsistency detected.\n"
            f"  File: {p.name}\n"
        )
        for e in errors:
            sys.stderr.write(f"  - {e}\n")
        sys.stderr.write(
            "\n"
            "  This usually means the Raw_Info or Assumptions tab was edited\n"
            "  without updating _State. Rebuild RAW_MAP/ASM_MAP before continuing.\n"
            "  See: ~/.claude/skills/3-statements-ultra/SKILL.md §Session startup Step 9\n"
        )
        sys.exit(2)

    sys.exit(0)


if __name__ == "__main__":
    main()
